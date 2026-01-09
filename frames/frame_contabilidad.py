import tkinter as tk
from tkinter import ttk, messagebox
from datetime import date, datetime, timedelta
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None  # type: ignore
from time import perf_counter as _pc
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

from frames.date_selector import DateSelector

from api_client.client import api_client, APIError
from resource_loader import asset_path

RED_BTN_BG = '#B18B83'
RED_BTN_ACTIVE = '#996D64'
BLACK_BTN_BG = '#111827'
BLACK_BTN_ACTIVE = '#1F2937'
CAFE_BG = '#B18B83'


class MultiSelectCombobox(ttk.Frame):
    def __init__(self, master, options, **kwargs):
        super().__init__(master, **kwargs)
        self.options = list(options)
        self.vars = {opt: tk.BooleanVar() for opt in self.options}
        # Solo mostrar el botón, sin etiqueta de selección
        self.button = ttk.Button(
            self,
            text="▼ Seleccionar",
            command=self.open_popup,
            style='Cafe.TButton'
        )
        self.button.pack(side='left', padx=5, pady=5)

    def open_popup(self):
        popup = tk.Toplevel(self)
        popup.title("Selecciona opciones")
        popup.transient(self)
        popup.grab_set()
        
        # Posicionar ventana justo debajo y alineada a la izquierda del botón
        button_x = self.button.winfo_rootx()
        button_y = self.button.winfo_rooty()
        popup_x = button_x
        popup_y = button_y + self.button.winfo_height() + 5
        popup.geometry(f"+{popup_x}+{popup_y}")
        
        # Frame para los checkbuttons
        check_frame = ttk.Frame(popup)
        check_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Agregar checks
        for opt in self.options:
            chk = ttk.Checkbutton(check_frame, text=opt, variable=self.vars[opt])
            chk.pack(anchor='w', padx=5, pady=2)
        
        def aceptar():
            self.update_selection(popup)
        ttk.Button(popup, text="Aceptar", command=aceptar).pack(pady=10)

    def update_selection(self, popup):
        selected = [opt for opt, var in self.vars.items() if var.get()]
        # Si selecciona "Todo", marcar todo
        if any(opt.lower() == 'todo' for opt in selected):
            for k in self.vars.values():
                k.set(True)
            selected = list(self.vars.keys())
        # No mostrar los items seleccionados, solo cerrar el popup
        popup.destroy()

    def get_selected(self):
        sel = [opt for opt, var in self.vars.items() if var.get()]
        return sel


class FrameContabilidad(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        # Timezone del token para alinear "hoy" con el backend
        try:
            tz = api_client.get_user_timezone()
            self.token_tz = tz or 'UTC'
        except Exception:
            self.token_tz = 'UTC'
        self._empleados = []
        self._ready = False
        # Cache de métricas y control de recálculo
        self._cached_metrics = None
        self._last_filters = None
        self._stale = True
        self._build_ui()
        # No cargar empleados ni métricas automáticamente
        self._ready = True

    def _build_ui(self):
        style = ttk.Style(self)
        try:
            style.theme_use('clam')
        except Exception:
            pass
        try:
            style.configure('Blue.TButton', font=('Segoe UI', 9, 'bold'), padding=(12, 8), background='#73D0E6')
            style.map('Blue.TButton', background=[('active', '#4FC3D9'), ('pressed', '#4FC3D9')])
            style.configure('Cafe.TButton', font=('Segoe UI', 9, 'bold'), padding=(12, 8), background=CAFE_BG, foreground='white')
            style.map('Cafe.TButton', background=[('active', '#996D64'), ('pressed', '#996D64')])
        except Exception:
            pass

        self.columnconfigure(1, weight=1)
        self.rowconfigure(1, weight=1)

        # Columna izquierda: rutas/empleados
        left = ttk.LabelFrame(self, text="Rutas")
        left.grid(row=0, column=0, rowspan=3, sticky='nsw', padx=6, pady=6)
        left.rowconfigure(1, weight=1)

        # Treeview sin columnas adicionales; se mostrará sobre una imagen de fondo
        self.tree = ttk.Treeview(left, selectmode='browse', show='tree', height=20)
        try:
            self.tree.column('#0', width=220, minwidth=180)
        except Exception:
            pass
        # Imagen de fondo opcional
        self._tree_bg = None
        try:
            from PIL import Image, ImageTk
            import os
            bg_path = asset_path('assets', 'logos', 'empleados_bg.png')
            if os.path.exists(bg_path):
                img = Image.open(bg_path)
                try:
                    resample = Image.Resampling.LANCZOS
                except Exception:
                    resample = Image.LANCZOS
                img = img.resize((240, 430), resample)
                self._tree_bg = ImageTk.PhotoImage(img)
                self.tree_bg_label = ttk.Label(left, image=self._tree_bg)
                self.tree_bg_label.place(x=2, y=46)
        except Exception:
            pass
        self.tree.grid(row=1, column=0, sticky='nsew', padx=4, pady=4)
        try:
            self.tree.lift()
        except Exception:
            pass

        # Flecha roja para empleados
        self._tree_arrow = tk.PhotoImage(width=14, height=14)
        arrow_color = '#8B0000'
        for dx in range(6):
            for dy in range(-dx, dx + 1):
                x = 2 + dx
                y = 6 + dy
                if 0 <= x < 14 and 0 <= y < 14:
                    self._tree_arrow.put(arrow_color, (x, y))
        for dy in range(-1, 2):
            self._tree_arrow.put(arrow_color, (1, 6 + dy))

        # Botón para cargar empleados bajo demanda (no bloquea inicio)
        self.btn_cargar_empleados = ttk.Button(
            left,
            text="Cargar empleados",
            command=self._load_empleados,
            style='Cafe.TButton'
        )
        self.btn_cargar_empleados.grid(row=0, column=0, sticky='ew', padx=4, pady=(4,0))
        # Insertar nodo Consolidados por defecto
        try:
            if not self.tree.get_children():
                self.tree.insert('', 'end', text='Consolidados', image=self._tree_arrow, open=True, values=('ALL',))
        except Exception:
            pass

        # Columna derecha: cabecera con periodo e información
        header = ttk.Frame(self)
        header.grid(row=0, column=1, sticky='ew', padx=6, pady=(6,0))
        header.columnconfigure(8, weight=1)

        ttk.Label(header, text="Periodo de tiempo:").grid(row=0, column=0, sticky='w')
        ttk.Label(header, text="Desde:").grid(row=0, column=1, sticky='e')
        self.desde = DateSelector(
            header,
            date_pattern='yyyy-MM-dd',
            width=12,
        )
        self.desde.grid(row=0, column=2, padx=(6,6), pady=2)
        ttk.Label(header, text="Hasta:").grid(row=0, column=3, sticky='e')
        self.hasta = DateSelector(
            header,
            date_pattern='yyyy-MM-dd',
            width=12,
        )
        self.hasta.grid(row=0, column=4, padx=(6,12), pady=2)
        self._set_default_dates()

        # Selector de información (popup con checkbuttons)
        ttk.Label(header, text="Información:").grid(row=0, column=5, sticky='w', padx=(6,0))
        metrics = [
            'Valor cobrado', 'Préstamos', 'Intereses', 'Gastos', 'Entradas', 'Base', 'Ganancia',
            'Salidas', 'Caja', 'Cartera en calle', 'Número de abonos', 'Efectivo Operativo', 'Total Clavos', 'Todo'
        ]
        self.metrics_combo = MultiSelectCombobox(header, metrics)
        self.metrics_combo.grid(row=0, column=6, padx=(6,0))

        self.btn_aplicar = ttk.Button(
            header,
            text="Aplicar",
            command=self._on_change_filters,
            style='Cafe.TButton'
        )
        self.btn_aplicar.grid(row=0, column=7, padx=8)

        # Secciones Datos y Manejo de caja
        body = ttk.Frame(self)
        body.grid(row=1, column=1, sticky='nsew', padx=6, pady=6)
        body.columnconfigure(0, weight=3)   # Datos
        body.columnconfigure(1, weight=3)   # Resumen
        body.columnconfigure(2, weight=2)   # Manejo de caja
        body.rowconfigure(0, weight=9)      # Fila principal (widgets)
        body.rowconfigure(1, weight=1)      # Fila inferior (botones de acciones)

        # Sección Datos (texto)
        datos_box = ttk.LabelFrame(body, text="Datos")
        datos_box.grid(row=0, column=0, sticky='nsew', padx=(0,6))
        datos_box.columnconfigure(0, weight=1)
        datos_box.rowconfigure(0, weight=1)
        
        datos_frame = tk.Frame(datos_box, bg='white')
        datos_frame.pack(fill='both', expand=True, padx=4, pady=4)
        
        # Grid para etiquetas alineadas
        datos_frame.columnconfigure(1, weight=1)
        
        label_font = ('Segoe UI', 9)
        val_font = ('Segoe UI', 9, 'bold')
        val_fg = '#1F2937'
        
        tk.Label(datos_frame, text="Total cobrado:", font=label_font, bg='white').grid(row=0, column=0, sticky='w', pady=2)
        self.lbl_cobrado = tk.Label(datos_frame, text="$ 0", font=val_font, fg=val_fg, bg='white')
        self.lbl_cobrado.grid(row=0, column=1, sticky='e', pady=2)
        
        tk.Label(datos_frame, text="Total bases asignadas:", font=label_font, bg='white').grid(row=1, column=0, sticky='w', pady=2)
        self.lbl_bases = tk.Label(datos_frame, text="$ 0", font=val_font, fg=val_fg, bg='white')
        self.lbl_bases.grid(row=1, column=1, sticky='e', pady=2)
        
        tk.Label(datos_frame, text="Total préstamos:", font=label_font, bg='white').grid(row=2, column=0, sticky='w', pady=2)
        self.lbl_prestamos = tk.Label(datos_frame, text="$ 0", font=val_font, fg=val_fg, bg='white')
        self.lbl_prestamos.grid(row=2, column=1, sticky='e', pady=2)
        
        tk.Label(datos_frame, text="Total gastos:", font=label_font, bg='white').grid(row=3, column=0, sticky='w', pady=2)
        self.lbl_gastos = tk.Label(datos_frame, text="$ 0", font=val_font, fg=val_fg, bg='white')
        self.lbl_gastos.grid(row=3, column=1, sticky='e', pady=2)
        
        tk.Label(datos_frame, text="Total salidas:", font=label_font, bg='white').grid(row=4, column=0, sticky='w', pady=2)
        self.lbl_salidas = tk.Label(datos_frame, text="$ 0", font=val_font, fg=val_fg, bg='white')
        self.lbl_salidas.grid(row=4, column=1, sticky='e', pady=2)
        
        tk.Label(datos_frame, text="Total entradas:", font=label_font, bg='white').grid(row=5, column=0, sticky='w', pady=2)
        self.lbl_entradas = tk.Label(datos_frame, text="$ 0", font=val_font, fg=val_fg, bg='white')
        self.lbl_entradas.grid(row=5, column=1, sticky='e', pady=2)
        
        tk.Label(datos_frame, text="Total intereses:", font=label_font, bg='white').grid(row=6, column=0, sticky='w', pady=2)
        self.lbl_intereses = tk.Label(datos_frame, text="$ 0", font=val_font, fg=val_fg, bg='white')
        self.lbl_intereses.grid(row=6, column=1, sticky='e', pady=2)

        tk.Label(datos_frame, text="Ganancia Neta:", font=label_font, bg='white').grid(row=7, column=0, sticky='w', pady=2)
        self.lbl_ganancia = tk.Label(datos_frame, text="$ 0", font=('Segoe UI', 9, 'bold'), fg='#059669', bg='white')
        self.lbl_ganancia.grid(row=7, column=1, sticky='e', pady=2)
        
        tk.Label(datos_frame, text="Cartera en la calle (Hasta):", font=label_font, bg='white').grid(row=8, column=0, sticky='w', pady=2)
        self.lbl_cartera = tk.Label(datos_frame, text="$ 0", font=val_font, fg=val_fg, bg='white')
        self.lbl_cartera.grid(row=8, column=1, sticky='e', pady=2)
        
        tk.Label(datos_frame, text="Cartera en la calle (Desde):", font=label_font, bg='white').grid(row=9, column=0, sticky='w', pady=2)
        self.lbl_cartera_desde = tk.Label(datos_frame, text="$ 0", font=val_font, fg=val_fg, bg='white')
        self.lbl_cartera_desde.grid(row=9, column=1, sticky='e', pady=2)

        # Nuevo Item: Efectivo (Cobrado + Base - Prestamos - Gastos)
        tk.Label(datos_frame, text="Efectivo Operativo:", font=label_font, bg='white').grid(row=10, column=0, sticky='w', pady=2)
        self.lbl_efectivo = tk.Label(datos_frame, text="$ 0", font=('Segoe UI', 10, 'bold'), fg='#059669', bg='white')
        self.lbl_efectivo.grid(row=10, column=1, sticky='e', pady=2)

        # Nuevo Item: Total Clavos (>60 días vencidos)
        tk.Label(datos_frame, text="Total Clavos (>60d):", font=label_font, bg='white').grid(row=11, column=0, sticky='w', pady=2)
        self.lbl_clavos = tk.Label(datos_frame, text="$ 0", font=('Segoe UI', 9, 'bold'), fg='#DC2626', bg='white')
        self.lbl_clavos.grid(row=11, column=1, sticky='e', pady=2)
        
        # Botón de limpieza (Escobita) - Más grande y posicionado fijo en el fondo
        self.btn_limpiar = tk.Button(datos_frame, text="🧹", font=('Segoe UI', 16), 
                                   command=self._limpiar_tableros, 
                                   relief='flat', bg='white', activebackground='#F3F4F6', bd=0)
        self.btn_limpiar.place(relx=0.0, rely=1.0, anchor='sw', x=2, y=-2)
        
        # Sección Resumen (espacio ampliado)
        resumen_box = ttk.LabelFrame(body, text="Resumen")
        resumen_box.grid(row=0, column=1, sticky='nsew', padx=(0,6))
        resumen_box.rowconfigure(0, weight=1)
        resumen_box.columnconfigure(0, weight=1)  # permitir expandir a lo ancho
        self.txt_resumen = tk.Text(resumen_box, height=10, width=40, wrap=tk.WORD)
        self.txt_resumen.grid(row=0, column=0, sticky='nsew', padx=4, pady=4)

        # Manejo de caja (derecha)
        caja_box = ttk.LabelFrame(body, text="Manejo de caja")
        caja_box.grid(row=0, column=2, sticky='nsew')
        caja_box.columnconfigure(1, weight=1)
        caja_box.columnconfigure(2, weight=1)

        ttk.Label(caja_box, text="Fecha:").grid(row=0, column=0, sticky='w', padx=6, pady=(6,2))
        self.caja_fecha = DateSelector(
            caja_box,
            date_pattern='yyyy-MM-dd',
            width=12,
        )
        self.caja_fecha.grid(row=0, column=1, columnspan=2, padx=6, pady=(6,2), sticky='ew')
        
        ttk.Label(caja_box, text="Caja existente:").grid(row=1, column=0, sticky='w', padx=6, pady=(4,2))
        self.caja_valor = ttk.Label(caja_box, text="$ 0")
        self.caja_valor.grid(row=1, column=1, sticky='w', pady=(4,2))
        self.btn_leer_caja = ttk.Button(
            caja_box,
            text="Leer",
            command=self._leer_caja,
            style='Cafe.TButton',
            width=10
        )
        self.btn_leer_caja.grid(row=1, column=2, padx=(4,6), pady=(4,2), sticky='e')

        ttk.Label(caja_box, text="Dar salida:").grid(row=2, column=0, sticky='w', padx=6, pady=(8,2))
        self.salida_val = ttk.Entry(caja_box, font=('Segoe UI', 10, 'bold'))
        self.salida_val.grid(row=2, column=1, columnspan=2, padx=6, pady=(8,2), sticky='ew')
        ttk.Label(caja_box, text="Observación:").grid(row=3, column=0, sticky='w', padx=6)
        self.salida_conc = ttk.Entry(caja_box)
        self.salida_conc.grid(row=3, column=1, columnspan=2, padx=6, pady=(0,6), sticky='ew')
        self.btn_registrar_salida = ttk.Button(
            caja_box,
            text="Registrar salida",
            command=self._registrar_salida,
            style='Cafe.TButton'
        )
        self.btn_registrar_salida.grid(row=4, column=0, columnspan=3, pady=6, padx=6, sticky='ew')

        # Separador o espacio
        ttk.Separator(caja_box, orient='horizontal').grid(row=5, column=0, columnspan=3, sticky='ew', padx=6, pady=4)

        # Sección Entrada
        ttk.Label(caja_box, text="Dar entrada:").grid(row=6, column=0, sticky='w', padx=6, pady=(4,2))
        self.entrada_val = ttk.Entry(caja_box, font=('Segoe UI', 10, 'bold'))
        self.entrada_val.grid(row=6, column=1, columnspan=2, padx=6, pady=(4,2), sticky='ew')
        ttk.Label(caja_box, text="Observación:").grid(row=7, column=0, sticky='w', padx=6)
        self.entrada_conc = ttk.Entry(caja_box)
        self.entrada_conc.grid(row=7, column=1, columnspan=2, padx=6, pady=(0,6), sticky='ew')
        self.btn_registrar_entrada = ttk.Button(
            caja_box,
            text="Registrar entrada",
            command=self._registrar_entrada,
            style='Blue.TButton'  # Diferente color para diferenciar
        )
        self.btn_registrar_entrada.grid(row=8, column=0, columnspan=3, pady=6, padx=6, sticky='ew')
        
        # Fila inferior de acciones (debajo de Datos y centrados)
        actions_frame = ttk.Frame(body)
        actions_frame.grid(row=1, column=0, columnspan=2, pady=(2,0), sticky='ew')
        actions_frame.columnconfigure(0, weight=1)
        actions_frame.columnconfigure(1, weight=1)
        actions_frame.columnconfigure(2, weight=1)

        btn_exportar = ttk.Button(
            actions_frame,
            text="Exportar a Excel",
            command=self._exportar_excel,
            style='Blue.TButton',
            width=18
        )
        btn_exportar.grid(row=0, column=0, padx=(0,6), pady=2, sticky='ew')
        
        btn_listar = ttk.Button(
            actions_frame,
            text="Listar Clientes",
            command=self._listar_clientes,
            style='Blue.TButton',
            width=18
        )
        btn_listar.grid(row=0, column=1, padx=(6,6), pady=2, sticky='ew')

        btn_informes = ttk.Button(
            actions_frame,
            text="Informes y gráficos",
            command=self._abrir_informes,
            style='Blue.TButton',
            width=20
        )
        btn_informes.grid(row=0, column=2, padx=(6,0), pady=2, sticky='ew')

        # Eventos
        for w in (self.desde, self.hasta):
            w.bind('<Return>', lambda e: (self._ready and self._on_change_filters()))

    def _hoy_local(self) -> date:
        try:
            if ZoneInfo is not None:
                return datetime.now(ZoneInfo(self.token_tz)).date()
        except Exception:
            pass
        return date.today()

    def _set_default_dates(self):
        hoy = self._hoy_local()
        try:
            self.desde.set_date(hoy - timedelta(days=7))
            self.hasta.set_date(hoy)
        except Exception:
            self.desde.delete(0, tk.END)
            self.desde.insert(0, (hoy - timedelta(days=7)).isoformat())
            self.hasta.delete(0, tk.END)
            self.hasta.insert(0, hoy.isoformat())
        # Caja: por defecto también debe ser "hoy" local
        try:
            self.caja_fecha.set_date(hoy)
        except Exception:
            pass

    def _load_empleados(self):
        try:
            # Asegurar nodo raíz
            roots = self.tree.get_children()
            if roots:
                root_id = roots[0]
            else:
                root_id = self.tree.insert('', 'end', text='Consolidados', open=True, values=('ALL',))
            # Limpiar hijos actuales excepto root
            for child in list(self.tree.get_children(root_id)):
                self.tree.delete(child)
            empleados = api_client.list_empleados()
            self._empleados = empleados or []
            for emp in (empleados or []):
                nombre = emp.get('nombre_completo', emp.get('identificacion',''))
                texto = f"➤ {nombre}"
                self.tree.insert(root_id, 'end', text=texto, image=self._tree_arrow, values=(emp.get('identificacion'),))
            # Seleccionar root por defecto
            self.tree.selection_set(root_id)
        except Exception:
            # No bloquear el UI si falla
            pass

    def _get_filters(self):
        try:
            desde = self.desde.get_date()
            hasta = self.hasta.get_date()
        except Exception:
            try:
                desde = datetime.strptime(self.desde.get().strip(), '%Y-%m-%d').date()
                hasta = datetime.strptime(self.hasta.get().strip(), '%Y-%m-%d').date()
            except Exception:
                messagebox.showerror('Error', 'Fechas inválidas. Use YYYY-MM-DD')
                return None
        if desde > hasta:
            messagebox.showerror('Error', "La fecha 'Desde' no puede ser mayor que 'Hasta'")
            return None
        sel = self.tree.selection()
        empleado_id = None
        if sel:
            val = self.tree.item(sel[0]).get('values')
            if val and val[0] != 'ALL':
                empleado_id = str(val[0])
        return {'desde': desde, 'hasta': hasta, 'empleado_id': empleado_id}

    def _on_change_filters(self):
        f = self._get_filters()
        if not f:
            return
        try:
            # Usar caché si los filtros no cambian y no está marcado como "stale"
            api_elapsed = 0.0
            if (self._last_filters == f) and (not self._stale) and self._cached_metrics:
                data = self._cached_metrics
            else:
                t0 = _pc()
                data = api_client.contabilidad_metricas(f['desde'], f['hasta'], f['empleado_id'])
                api_elapsed = _pc() - t0
                # Actualizar caché y estado
                self._cached_metrics = data
                self._last_filters = dict(f)
                self._stale = False
                # Debug temporal: ver claves y valores críticos

            # Mostrar según selección múltiple
            seleccion = self.metrics_combo.get_selected()
            if not seleccion or ('Todo' in seleccion):
                seleccion = ['Valor cobrado','Préstamos','Intereses','Gastos','Entradas','Base','Ganancia','Salidas','Caja','Cartera en calle','Número de abonos','Efectivo Operativo','Total Clavos']
            self._render_outputs(data, seleccion)
            # Si se necesita cartera en calle y el backend dio 0, disparar cálculo async sin bloquear (con deduplicación)
            if ('Cartera en calle' in seleccion) and f.get('empleado_id') and float(data.get('cartera_en_calle', 0) or 0) == 0.0:
                key = (f.get('empleado_id'), f.get('hasta'))
                if getattr(self, '_cartera_fb_key', None) != key:
                    self._cartera_fb_key = key
                    threading.Thread(target=self._compute_cartera_fallback_async, args=(key[0], key[1]), daemon=True).start()
            # Perf por item (nota: casi todos salen de la misma llamada)
            for etiqueta in seleccion:
                pass
        except APIError as e:
            # Si el backend aún no tiene el endpoint, intentar un cálculo de respaldo
            try:
                if getattr(e, 'status_code', None) == 404:
                    t0 = _pc()
                    data = self._calc_metrics_fallback(f['desde'], f['hasta'], f['empleado_id'])
                    fb_elapsed = _pc() - t0
                    try:
                        seleccion = self.metrics_combo.get_selected()
                    except Exception:
                        seleccion = []
                    self._render_resumen_financiero(data, seleccion)
                    return
            except Exception:
                pass
            messagebox.showerror('Error', f"No se pudo obtener datos: {e.message}")
        except Exception as e:
            messagebox.showerror('Error', f"No se pudo obtener datos: {e}")

    def _registrar_salida(self):
        try:
            fecha = datetime.strptime(self.caja_fecha.get().strip(), '%Y-%m-%d').date()
        except Exception:
            messagebox.showerror('Error', 'Fecha inválida (YYYY-MM-DD)')
            return
        # Parseo robusto de monto (acepta 10.000, 10,50, 10000.50)
        try:
            raw = self.salida_val.get().strip()
            valor = self._parse_money_input(raw)
        except Exception:
            messagebox.showerror('Error', 'Valor inválido')
            return
        if valor is None or valor <= 0:
            messagebox.showerror('Error', 'El valor debe ser mayor a 0')
            return
        # Mostrar formateado en el campo
        try:
            self.salida_val.delete(0, tk.END)
            self.salida_val.insert(0, f"$ {valor:,.2f}")
        except Exception:
            pass
        try:
            sel = self.tree.selection()
            empleado_id = None
            if sel:
                val = self.tree.item(sel[0]).get('values')
                if val and val[0] != 'ALL':
                    empleado_id = str(val[0])
            if not empleado_id:
                messagebox.showerror('Error', 'Seleccione un empleado para registrar una salida de caja')
                return
            _ = api_client.registrar_salida_caja(
                fecha=fecha,
                valor=valor,
                concepto=(self.salida_conc.get().strip() or None),
                empleado_identificacion=empleado_id,
            )
            messagebox.showinfo('OK', 'Salida registrada')
            # Marcar como pendiente de recálculo; el usuario decidirá cuándo aplicar
            self._stale = True
        except APIError as e:
            messagebox.showerror('Error', f"No se pudo registrar la salida: {e.message}")
        except Exception as e:
            messagebox.showerror('Error', f"No se pudo registrar la salida: {e}")

    def _registrar_entrada(self):
        try:
            fecha = datetime.strptime(self.caja_fecha.get().strip(), '%Y-%m-%d').date()
        except Exception:
            messagebox.showerror('Error', 'Fecha inválida (YYYY-MM-DD)')
            return
        # Parseo robusto de monto
        try:
            raw = self.entrada_val.get().strip()
            valor = self._parse_money_input(raw)
        except Exception:
            messagebox.showerror('Error', 'Valor inválido')
            return
        if valor is None or valor <= 0:
            messagebox.showerror('Error', 'El valor debe ser mayor a 0')
            return
        # Mostrar formateado en el campo
        try:
            self.entrada_val.delete(0, tk.END)
            self.entrada_val.insert(0, f"$ {valor:,.2f}")
        except Exception:
            pass
        try:
            sel = self.tree.selection()
            empleado_id = None
            if sel:
                val = self.tree.item(sel[0]).get('values')
                if val and val[0] != 'ALL':
                    empleado_id = str(val[0])
            if not empleado_id:
                messagebox.showerror('Error', 'Seleccione un empleado para registrar una entrada de caja')
                return
            _ = api_client.registrar_entrada_caja(
                fecha=fecha,
                valor=valor,
                concepto=(self.entrada_conc.get().strip() or None),
                empleado_identificacion=empleado_id,
            )
            messagebox.showinfo('OK', 'Entrada registrada')
            # Marcar como pendiente de recálculo
            self._stale = True
        except APIError as e:
            messagebox.showerror('Error', f"No se pudo registrar la entrada: {e.message}")
        except Exception as e:
            messagebox.showerror('Error', f"No se pudo registrar la entrada: {e}")

    def _leer_caja(self):
        try:
            # Empleado seleccionado
            sel = self.tree.selection()
            empleado_id = None
            if sel:
                val = self.tree.item(sel[0]).get('values')
                if val and val[0] != 'ALL':
                    empleado_id = str(val[0])
            if not empleado_id:
                messagebox.showwarning('Aviso', 'Seleccione un empleado para leer caja')
                return
            # Fecha
            try:
                fecha = datetime.strptime(self.caja_fecha.get().strip(), '%Y-%m-%d').date()
            except Exception:
                messagebox.showerror('Error', 'Fecha inválida (YYYY-MM-DD)')
                return
            # Llamar API y mostrar
            data = api_client.caja_valor(empleado_id, fecha)
            val = float(data.get('valor', 0) or 0)
            self.caja_valor.configure(text=f"$ {val:,.2f}")
        except APIError as e:
            messagebox.showerror('Error', f"No se pudo leer caja: {e.message}")
        except Exception as e:
            messagebox.showerror('Error', f"No se pudo leer caja: {e}")

    def _calc_metrics_fallback(self, desde: date, hasta: date, empleado_id: str | None) -> dict:
        """Suma liquidaciones diarias existentes como respaldo.
        Salidas=0 (aún no consolidadas por rango en fallback).
        """
        total_cobrado = 0.0
        total_prestamos = 0.0
        total_gastos = 0.0
        total_bases = 0.0
        empleados = []
        if empleado_id:
            empleados = [empleado_id]
        else:
            try:
                if not self._empleados:
                    self._empleados = api_client.list_empleados()
                empleados = [e.get('identificacion') for e in self._empleados if e.get('identificacion')]
            except Exception:
                empleados = []

        d = desde
        step = timedelta(days=1)
        while d <= hasta:
            for emp in empleados:
                try:
                    liq = api_client.get_liquidacion_diaria(emp, d)
                    total_cobrado += float(liq.get('total_recaudado', 0) or 0)
                    total_prestamos += float(liq.get('prestamos_otorgados', 0) or 0)
                    total_gastos += float(liq.get('total_gastos', 0) or 0)
                    total_bases += float(liq.get('base_dia', 0) or 0)
                except Exception:
                    pass
            d += step

        total_salidas = 0.0
        total_entradas = 0.0
        # Fallback simple sin entradas (requeriría nueva API) y SIN BASE (nueva fórmula)
        caja_calculada = total_cobrado - total_prestamos - total_gastos - total_salidas + total_entradas
        return {
            'total_cobrado': total_cobrado,
            'total_prestamos': total_prestamos,
            'total_gastos': total_gastos,
            'total_bases': total_bases,
            'total_salidas': total_salidas,
            'total_entradas': total_entradas,
            # En fallback devolvemos 'caja' para compatibilidad con la UI
            'caja': caja_calculada,
        }

    def _listar_clientes(self):
        # Verificar selección de empleado
        sel = self.tree.selection()
        empleado_id = None
        if sel:
            val = self.tree.item(sel[0]).get('values')
            if val and val[0] != 'ALL':
                empleado_id = str(val[0])
        
        if not empleado_id:
            messagebox.showwarning("Selección", "Por favor seleccione un empleado específico de la lista.")
            return

        # Diálogo para elegir alcance
        top = tk.Toplevel(self)
        top.title("Listar Clientes")
        top.geometry("320x160")
        top.resizable(False, False)
        top.transient(self)
        try:
            top.grab_set()
            # Centrar
            top.update_idletasks()
            x = self.winfo_rootx() + (self.winfo_width() // 2) - 160
            y = self.winfo_rooty() + (self.winfo_height() // 2) - 80
            top.geometry(f"+{x}+{y}")
        except:
            pass

        tk.Label(top, text="¿Qué clientes desea listar?", font=('Segoe UI', 11, 'bold')).pack(pady=(20, 10))
        
        def action(scope):
            top.destroy()
            self._generar_excel_clientes(empleado_id, scope)

        f_btns = tk.Frame(top)
        f_btns.pack(fill='x', padx=20, pady=10)
        
        ttk.Button(f_btns, text="Solo Activos\n(con deuda)", style='Blue.TButton', 
                   command=lambda: action('activos')).pack(side='left', expand=True, padx=5, fill='x')
        ttk.Button(f_btns, text="Todos\n(histórico)", style='Cafe.TButton', 
                   command=lambda: action('todos')).pack(side='left', expand=True, padx=5, fill='x')

    def _generar_excel_clientes(self, empleado_id: str, scope: str):
        # Lógica similar a _exportar_excel pero simplificada para clientes
        try:
            import os, sys, tempfile
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Ventana de progreso
            prog = tk.Toplevel(self)
            prog.title('Generando lista...')
            prog.geometry("300x120")
            prog.resizable(False, False)
            try:
                prog.transient(self)
                prog.grab_set()
                prog.update_idletasks()
                x = self.winfo_rootx() + (self.winfo_width() // 2) - 150
                y = self.winfo_rooty() + (self.winfo_height() // 2) - 60
                prog.geometry(f"+{x}+{y}")
            except: pass
            
            ttk.Label(prog, text="Consultando y generando Excel...").pack(pady=20)
            pbar = ttk.Progressbar(prog, mode='indeterminate')
            pbar.pack(fill='x', padx=20)
            pbar.start(10)
            
            def _worker():
                try:
                    # 1. Fetch data
                    clientes = api_client.list_clientes_empleado(empleado_id, scope)
                    
                    if not clientes:
                        raise ValueError("No se encontraron clientes para este criterio.")

                    # 2. Build Excel
                    wb = Workbook()
                    ws = wb.active
                    ws.title = f"Clientes - {scope.capitalize()}"
                    
                    cols = ['Identificación', 'Nombre', 'Apellido', 'Teléfono', 'Dirección']
                    ws.append(cols)
                    
                    # Estilos
                    header_fill = PatternFill(start_color='FF0B63B5', end_color='FF0B63B5', fill_type='solid')
                    header_font = Font(color='FFFFFFFF', bold=True)
                    
                    for c_idx, _ in enumerate(cols, 1):
                        cell = ws.cell(row=1, column=c_idx)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                        
                    for c in clientes:
                        ws.append([
                            c.get('identificacion'),
                            c.get('nombre'),
                            c.get('apellido'),
                            c.get('telefono') or '',
                            c.get('direccion') or ''
                        ])
                        
                    # Auto-width simple
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except: pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column].width = min(adjusted_width, 50)

                    tmp_path = os.path.join(tempfile.gettempdir(), f"clientes_{empleado_id}_{scope}.xlsx")
                    wb.save(tmp_path)
                    
                    def _finish():
                        try:
                            pbar.stop()
                            prog.destroy()
                        except: pass
                        
                        # Abrir archivo
                        try:
                            if sys.platform.startswith('win'):
                                os.startfile(tmp_path)
                            else:
                                import webbrowser
                                webbrowser.open(tmp_path)
                        except Exception as e:
                            messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")

                    self.after(0, _finish)
                    
                except Exception as e:
                    error_msg = str(e)
                    def _err(msg=error_msg):
                        try: pbar.stop(); prog.destroy() 
                        except: pass
                        messagebox.showerror("Error", f"Fallo al generar lista: {msg}")
                    self.after(0, _err)

            import threading
            threading.Thread(target=_worker, daemon=True).start()

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo iniciar: {e}")

    def _limpiar_tableros(self):
        """Limpia los datos mostrados y el caché."""
        self._cached_metrics = None
        self._stale = True
        
        # Limpiar labels de datos
        for lbl in [self.lbl_cobrado, self.lbl_bases, self.lbl_prestamos, 
                   self.lbl_gastos, self.lbl_salidas, self.lbl_entradas, 
                   self.lbl_intereses, self.lbl_cartera, self.lbl_cartera_desde,
                   self.lbl_efectivo, getattr(self, 'lbl_clavos', None),
                   getattr(self, 'lbl_ganancia', None)]:
            if lbl:
                lbl.config(text="$ 0")
        
        # Limpiar resumen texto
        self.txt_resumen.delete('1.0', tk.END)
        self.txt_resumen.insert('1.0', "Datos limpiados.")

    def _exportar_excel(self):
        # Previsualización en Excel (xlsx) con estilos, resúmenes y ventana de progreso centrada.
        f = self._get_filters()
        if not f:
            return
        try:
            import tkinter as tk
            from tkinter import ttk, messagebox
            import os, sys, subprocess, tempfile, time
            from datetime import timedelta
            from concurrent.futures import ThreadPoolExecutor, as_completed
            from time import perf_counter as _pc

            # Ventana de progreso modal (centrada)
            prog = tk.Toplevel(self)
            prog.title('Creando previsualización...')
            prog.resizable(False, False)
            prog.transient(self.winfo_toplevel())
            try:
                prog.grab_set()
            except Exception:
                pass
            # Estilo verde para la barra indeterminada
            try:
                style = ttk.Style(prog)
                style.configure('Success.Horizontal.TProgressbar', background='#10B981')
            except Exception:
                style = None
            msg = ttk.Label(prog, text='Generando archivo de Excel temporal, por favor espere...')
            msg.pack(padx=16, pady=(16, 8))
            pbar = ttk.Progressbar(prog, mode='indeterminate', length=280, style='Success.Horizontal.TProgressbar')
            pbar.pack(padx=16, pady=(0, 16))
            prog.update_idletasks()
            w, h = 360, 140
            x = (prog.winfo_screenwidth() // 2) - (w // 2)
            y = (prog.winfo_screenheight() // 2) - (h // 2)
            prog.geometry(f"{w}x{h}+{x}+{y}")
            pbar.start(12)
            self.update_idletasks()

            def _open_path(p: str):
                try:
                    if sys.platform.startswith('win'):
                        os.startfile(p)  # type: ignore[attr-defined]
                    elif sys.platform == 'darwin':
                        subprocess.Popen(['open', p])
                    else:
                        subprocess.Popen(['xdg-open', p])
                except Exception:
                    try:
                        import webbrowser
                        webbrowser.open(p)
                    except Exception:
                        pass

            def _worker():
                t0 = _pc()
                try:
                    try:
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, PatternFill, Alignment
                        from openpyxl.utils import get_column_letter
                    except Exception as e:
                        raise RuntimeError('Falta la dependencia openpyxl. Instale con: pip install openpyxl') from e

                    # Preparar fechas diarias
                    start = f['desde']; end = f['hasta']; empleado_id = f['empleado_id']
                    fechas = []
                    d = start
                    step = timedelta(days=1)
                    while d <= end:
                        fechas.append(d)
                        d += step

                    # Obtener métricas diarias en paralelo
                    t_fetch0 = _pc()
                    daily_rows = [None] * len(fechas)
                    max_workers = min(8, max(1, len(fechas)))

                    def fetch_one(idx: int, dia):
                        data = api_client.contabilidad_metricas(dia, dia, empleado_id)
                        # Usar el dato histórico calculado por el backend
                        total_posibles = int(data.get('tarjetas_activas_historicas', 0))
                        
                        # Fallback a liquidación guardada si el backend retorna 0
                        if total_posibles == 0:
                            try:
                                liq = api_client.get_liquidacion_diaria(empleado_id, dia)
                                total_posibles = int(liq.get('tarjetas_activas', 0))
                            except:
                                pass
                            
                        abonos_hechos = int(data.get('abonos_count', 0))
                        abonos_str = f"{abonos_hechos} de {total_posibles} posibles" if total_posibles > 0 else f"{abonos_hechos}"

                        # Efectivo (Cobrado + Base - Prestamos - Gastos)
                        efectivo = float(data.get('total_efectivo', 0))

                        return idx, {
                            'Fecha': dia.isoformat(),
                            'Número de abonos': abonos_str, # String ahora
                            'abonos_val': abonos_hechos,    # Para sumas numéricas
                            'posibles_val': total_posibles, # Para sumas numéricas
                            'Valor cobrado': float(data.get('total_cobrado', 0)),
                            'Préstamos': float(data.get('total_prestamos', 0)),
                            'Intereses': float(data.get('total_intereses', 0)),
                            'Bases': float(data.get('total_bases', 0)),
                            'Gastos': float(data.get('total_gastos', 0)),
                            'Ganancias': float(data.get('ganancia', 0)),
                            'Salidas': float(data.get('total_salidas', 0)),
                            'Entradas': float(data.get('total_entradas', 0) or 0), # Asegurar valor no nulo
                            'Efectivo': efectivo,
                            'Caja': float(data.get('caja', 0)),
                            'Cartera en calle': float(data.get('cartera_en_calle', 0)),
                        }

                    with ThreadPoolExecutor(max_workers=max_workers) as ex:
                        futures = [ex.submit(fetch_one, idx, dia) for idx, dia in enumerate(fechas)]
                        for fut in as_completed(futures):
                            idx, row = fut.result()
                            daily_rows[idx] = row

                    t_fetch = _pc() - t_fetch0

                    # Construir filas finales con semanas y total mes
                    # Orden de columnas solicitado: ..., Entradas, Efectivo, Caja, ...
                    cols = ['Fecha','Número de abonos','Valor cobrado','Préstamos','Intereses','Bases','Gastos','Ganancias','Salidas','Entradas','Efectivo','Caja','Cartera en calle']
                    
                    # Keys numéricas para sumar
                    sum_keys = ['abonos_val','posibles_val','Valor cobrado','Préstamos','Intereses','Bases','Gastos','Ganancias','Salidas','Entradas','Efectivo']
                    last_keys = ['Caja','Cartera en calle']

                    def _zero_accum():
                        return {k: 0.0 for k in sum_keys}

                    def _append_summary(rows_out, label, sums, last_vals, style_tag=None):
                        row = {'Fecha': label}
                        # Formatear abonos combinados
                        abonos = int(sums.get('abonos_val', 0))
                        posibles = int(sums.get('posibles_val', 0))
                        row['Número de abonos'] = f"{abonos} de {posibles} posibles"
                        
                        for k in sum_keys:
                            if k not in ('abonos_val', 'posibles_val'):
                                row[k] = float(sums.get(k, 0.0))
                        for lk in last_keys:
                            row[lk] = float(last_vals.get(lk, 0.0))
                        
                        row['_style'] = style_tag
                        rows_out.append(row)

                    final_rows = []
                    week_idx = 1
                    days_in_week = 0
                    weeks_in_month_block = 0
                    week_sums = _zero_accum(); week_last = {k: 0.0 for k in last_keys}
                    month_sums = _zero_accum(); month_last = {k: 0.0 for k in last_keys}
                    
                    # Acumuladores para total intervalo
                    total_sums = _zero_accum()
                    total_last = {k: 0.0 for k in last_keys}

                    total_days = len(daily_rows)
                    for i, r in enumerate(daily_rows, start=1):
                        final_rows.append(r)
                        for k in sum_keys:
                            val = float(r.get(k, 0) or 0)
                            week_sums[k] += val
                            total_sums[k] += val
                        for lk in last_keys:
                            val = float(r.get(lk, 0) or 0)
                            week_last[lk] = val
                            total_last[lk] = val # Sobrescribir con el último valor
                        
                        days_in_week += 1

                        is_last_day = (i == total_days)
                        if days_in_week == 7 or is_last_day:
                            _append_summary(final_rows, f"Semana {week_idx}", week_sums, week_last, 'week')
                            for k in sum_keys:
                                month_sums[k] += week_sums[k]
                            for lk in last_keys:
                                month_last[lk] = week_last[lk]
                            days_in_week = 0
                            week_idx += 1
                            weeks_in_month_block += 1
                            week_sums = _zero_accum(); week_last = {k: 0.0 for k in last_keys}
                        if weeks_in_month_block == 4 or (is_last_day and weeks_in_month_block > 0):
                            _append_summary(final_rows, 'Total mes', month_sums, month_last, 'month')
                            weeks_in_month_block = 0
                            month_sums = _zero_accum(); month_last = {k: 0.0 for k in last_keys}
                    
                    # Total general del intervalo
                    _append_summary(final_rows, 'TOTAL INTERVALO', total_sums, total_last, 'total')

                    # Crear libro Excel con estilos
                    t_build0 = _pc()
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils import get_column_letter
                    wb = Workbook(); ws = wb.active; ws.title = 'Contabilidad'
                    
                    # Encabezado fijo
                    ws.freeze_panes = 'A2'

                    # Estilos
                    header_fill = PatternFill(start_color='FFDC2626', end_color='FFDC2626', fill_type='solid') # Rojo como antes
                    header_font = Font(color='FFFFFFFF', bold=True)
                    
                    # Colores de filas de resumen
                    week_fill = PatternFill(start_color='FFFCD34D', end_color='FFFCD34D', fill_type='solid') # Ocre/Amarillo suave
                    month_fill = PatternFill(start_color='FFF59E0B', end_color='FFF59E0B', fill_type='solid') # Amarillo intenso
                    total_fill = PatternFill(start_color='FFBAE6FD', end_color='FFBAE6FD', fill_type='solid') # Azul claro
                    month_fill = PatternFill(start_color='FFF59E0B', end_color='FFF59E0B', fill_type='solid')  # ámbar

                    # Encabezados
                    ws.append(cols)
                    for c_idx in range(1, len(cols) + 1):
                        cell = ws.cell(row=1, column=c_idx)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Filas
                    for r_data in final_rows:
                        row_vals = []
                        for c in cols:
                            row_vals.append(r_data.get(c))
                        
                        ws.append(row_vals)
                        curr_row = ws.max_row
                        
                        # Formatos numéricos
                        for c_idx in range(2, len(cols) + 1):
                            key = cols[c_idx-1]
                            cell = ws.cell(row=curr_row, column=c_idx)
                            if key == 'Número de abonos':
                                cell.number_format = '@' # Texto para "X de Y"
                            else:
                                cell.number_format = '#,##0.00'

                        # Detectar filas de resumen por tag
                        style_tag = r_data.get('_style')
                        val_fecha = str(r_data.get('Fecha', ''))
                        
                        fill = None
                        if style_tag == 'week' or 'Semana' in val_fecha: 
                            fill = week_fill
                        elif style_tag == 'month' or 'Total mes' in val_fecha: 
                            fill = month_fill
                        elif style_tag == 'total' or 'TOTAL' in val_fecha: 
                            fill = total_fill
                        
                        if fill:
                            for c_idx in range(1, len(cols) + 1):
                                cell = ws.cell(row=curr_row, column=c_idx)
                                cell.fill = fill
                                cell.font = Font(bold=True)

                    # Ancho de columnas: más espaciosas (ajustadas)
                    width_map = {
                        'Fecha': 16,
                        'Número de abonos': 20,
                        'Valor cobrado': 20,
                        'Préstamos': 20,
                        'Intereses': 20,
                        'Bases': 20,
                        'Gastos': 20,
                        'Ganancias': 20,
                        'Salidas': 20,
                        'Entradas': 20,
                        'Efectivo': 20,
                        'Caja': 20,
                        'Cartera en calle': 20,
                    }
                    for col_idx, key in enumerate(cols, start=1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = width_map.get(key, 18)

                    # Guardar en archivo temporal y abrir (previsualización)
                    tmp_path = os.path.join(tempfile.gettempdir(), f"contabilidad_preview_{os.getpid()}_{int(time.time())}.xlsx")
                    wb.save(tmp_path)
                    t_build = _pc() - t_build0
                    total = _pc() - t0

                    def _done_ok():
                        try:
                            pbar.stop(); prog.destroy()
                        except Exception:
                            pass
                        _open_path(tmp_path)
                        # Ofrecer "Guardar como..." después de abrir la previsualización
                        try:
                            from tkinter import filedialog, messagebox as _mb
                            desea_guardar = _mb.askyesno('Guardar', '¿Desea guardar una copia del reporte?')
                            if desea_guardar:
                                save_path = filedialog.asksaveasfilename(
                                    defaultextension='.xlsx',
                                    filetypes=[('Excel Workbook','*.xlsx')],
                                    title='Guardar copia del reporte en Excel'
                                )
                                if save_path:
                                    import shutil as _sh
                                    _sh.copyfile(tmp_path, save_path)
                        except Exception:
                            pass
                        try:
                            from logging import getLogger
                            getLogger(__name__).info(f"[perf][conta][excel_preview] dias={len(fechas)} fetch={t_fetch:.3f}s build={t_build:.3f}s total={total:.3f}s workers={max_workers}")
                        except Exception:
                            pass
                    try:
                        self.after(0, _done_ok)
                    except Exception:
                        _done_ok()
                except Exception as e:
                    error_msg = str(e) # Capturar el mensaje inmediatamente
                    def _done_err(err=error_msg): # Usar argumento por defecto o variable capturada
                        try:
                            pbar.stop(); prog.destroy()
                        except Exception:
                            pass
                        messagebox.showerror('Error', f'No se pudo generar la previsualización de Excel: {err}')
                    try:
                        self.after(0, _done_err)
                    except Exception:
                        _done_err()

            # Ejecutar en segundo plano
            import threading
            threading.Thread(target=_worker, daemon=True).start()
        except Exception as e:
            try:
                from tkinter import messagebox
                messagebox.showerror('Error', f'No se pudo iniciar la exportación: {e}')
            except Exception:
                pass

    def _abrir_informes(self):
        messagebox.showinfo('Informes', 'Función futura: abrir informes y gráficos web')

    # API del contenedor
    def refrescar_datos(self):
        # No recalcular automáticamente al volver al frame
        pass

    # Utilidades internas
    def _parse_money_input(self, raw: str) -> float | None:
        s = (raw or '').strip()
        if not s:
            return None
        # Normalizar formato: quitar separadores de miles y usar '.' decimal
        s = s.replace('.', '').replace(',', '.')
        return float(s)

    def _fallback_cartera_en_calle(self, empleado_id: str, fecha_corte: date) -> float | None:
        """Suma el saldo pendiente de todas las tarjetas activas del empleado.
        Usado como respaldo si la API devuelve 0.
        """
        try:
            tarjetas = api_client.list_tarjetas(empleado_id=empleado_id, estado='activas', skip=0, limit=5000) or []
            n = len(tarjetas)
            # Evitar operaciones gigantes
            if n == 0:
                return 0.0
            if n > 1200:
                return None
            total = 0.0
            # Paralelizar consultas de resumen para acelerar
            max_workers = 8
            def fetch_saldo(codigo: str) -> float:
                try:
                    resumen = api_client.get_tarjeta_resumen(codigo)
                    return float(resumen.get('saldo_pendiente', 0) or 0)
                except Exception:
                    return 0.0
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = []
                for t in tarjetas:
                    codigo = t.get('codigo')
                    if not codigo:
                        continue
                    futures.append(executor.submit(fetch_saldo, codigo))
                for fut in as_completed(futures):
                    saldo = fut.result()
                    if saldo > 0:
                        total += saldo
            return total
        except Exception:
            return None

    def _render_outputs(self, data: dict, seleccion: list[str]) -> None:
        # Render de Resumen (incluye actualización de labels de datos)
        self.txt_resumen.delete('1.0', tk.END)
        self._render_resumen_financiero(data, seleccion)

    def _render_resumen_financiero(self, data: dict, seleccion: list[str]) -> None:
        """Muestra un resumen didáctico con ecuaciones usando SOLO datos ya cargados.
        Si falta algún operando por no estar en la selección, marcar como 'no solicitado'.
        """
        # Actualizar labels de métricas (panel izquierdo)
        try:
            self.lbl_cobrado.config(text=f"$ {data.get('total_cobrado', 0):,.0f}")
            self.lbl_bases.config(text=f"$ {data.get('total_bases', 0):,.0f}")
            self.lbl_prestamos.config(text=f"$ {data.get('total_prestamos', 0):,.0f}")
            self.lbl_gastos.config(text=f"$ {data.get('total_gastos', 0):,.0f}")
            self.lbl_salidas.config(text=f"$ {data.get('total_salidas', 0):,.0f}")
            self.lbl_entradas.config(text=f"$ {data.get('total_entradas', 0):,.0f}")
            self.lbl_intereses.config(text=f"$ {data.get('total_intereses', 0):,.0f}")
            self.lbl_cartera.config(text=f"$ {data.get('cartera_en_calle', 0):,.0f}")
            self.lbl_cartera_desde.config(text=f"$ {data.get('cartera_en_calle_desde', 0):,.0f}")
            self.lbl_efectivo.config(text=f"$ {data.get('total_efectivo', 0):,.0f}")
            self.lbl_clavos.config(text=f"$ {data.get('total_clavos', 0):,.0f}")
        except Exception:
            pass

        # Configurar estilos de texto
        try:
            self.txt_resumen.tag_config('title', font=('Segoe UI', 11, 'bold'))
            self.txt_resumen.tag_config('utilidad', foreground='#2563EB', font=('Segoe UI', 10, 'bold'))
            self.txt_resumen.tag_config('caja', foreground='#059669', font=('Segoe UI', 10, 'bold'))
            self.txt_resumen.tag_config('cartera', foreground='#F59E0B', font=('Segoe UI', 10, 'bold'))
            self.txt_resumen.tag_config('mono', font=('Consolas', 10))
            self.txt_resumen.tag_config('muted', foreground='#6B7280')
        except Exception:
            pass

        def fmt_money(v: float | int | None) -> str:
            try:
                return f"$ {float(v or 0):,.0f}"
            except Exception:
                return "$ 0"

        sel = set(seleccion or [])
        if ('Todo' in sel) or (not sel):
            # Considerar todo habilitado si selección vacía o 'Todo'
            sel = {'Valor cobrado','Préstamos','Intereses','Gastos','Base','Entradas','Salidas','Caja','Cartera en calle','Número de abonos'}

        # Operandos requeridos por cada ecuación
        need_utilidad = ('Intereses', 'Gastos')
        # Caja: Cobrado - Prestamos - Gastos - Salidas + Entradas (Base eliminada)
        need_caja = ('Valor cobrado', 'Préstamos', 'Gastos', 'Salidas', 'Entradas')
        # Crecimiento de cartera requiere dos puntos (desde/hasta) → no forzamos llamadas adicionales
        # Usamos solo 'Cartera en calle' si existiese par (desde/hasta) en data; si no, marcamos no solicitado

        def have_all(labels: tuple[str, ...]) -> bool:
            return all(lbl in sel for lbl in labels)

        # Limpiar
        self.txt_resumen.delete('1.0', tk.END)

        # Título
        self.txt_resumen.insert('end', 'RESUMEN FINANCIERO\n', ('title',))
        self.txt_resumen.insert('end', '\n')

        # 1) Utilidad del periodo = Intereses – Gastos
        i_val = data.get('total_intereses') if 'Intereses' in sel else None
        g_val = data.get('total_gastos') if 'Gastos' in sel else None
        util_val = (float(i_val) - float(g_val)) if (i_val is not None and g_val is not None) else None
        self.txt_resumen.insert('end', 'Utilidad del periodo: ', ('utilidad',))
        self.txt_resumen.insert('end', '\n')
        self.txt_resumen.insert('end', '  Intereses ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(i_val) if i_val is not None else ' (no solicitado)', (None if i_val is not None else 'muted',))
        self.txt_resumen.insert('end', '  −  Gastos ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(g_val) if g_val is not None else ' (no solicitado)', (None if g_val is not None else 'muted',))
        self.txt_resumen.insert('end', '  =  ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(util_val) if util_val is not None else ' —', ('utilidad',))
        self.txt_resumen.insert('end', '\n\n')

        # 2) Caja del periodo = Cobrado - Préstamos - Gastos - Salidas + Entradas
        c_val = data.get('total_cobrado') if 'Valor cobrado' in sel else None
        # b_val = data.get('total_bases') if 'Base' in sel else None  <-- Base eliminada de ecuación caja
        p_val = data.get('total_prestamos') if 'Préstamos' in sel else None
        g2_val = data.get('total_gastos') if 'Gastos' in sel else None
        s_val = data.get('total_salidas') if 'Salidas' in sel else None
        e_val = data.get('total_entradas') if 'Entradas' in sel else None
        
        # Asegurar que si Entradas no está en metrics, se trate como 0 si no fue solicitada, o None si era requerida
        # Pero para consistencia, si el usuario no pidió Entradas, no mostramos la ecuación completa o mostramos parcial
        # Asumiremos que Entradas es parte esencial ahora.
        
        caja_calc = (float(c_val) - float(p_val) - float(g2_val) - float(s_val) + float(e_val)) if all(v is not None for v in (c_val,p_val,g2_val,s_val,e_val)) else None
        
        self.txt_resumen.insert('end', 'Caja del periodo: ', ('caja',))
        self.txt_resumen.insert('end', '\n')
        self.txt_resumen.insert('end', '  Cobrado ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(c_val) if c_val is not None else ' (no solicitado)', (None if c_val is not None else 'muted',))
        self.txt_resumen.insert('end', '  +  Entradas ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(e_val) if e_val is not None else ' (no solicitado)', (None if e_val is not None else 'muted',))
        self.txt_resumen.insert('end', '  −  Préstamos ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(p_val) if p_val is not None else ' (no solicitado)', (None if p_val is not None else 'muted',))
        self.txt_resumen.insert('end', '  −  Gastos ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(g2_val) if g2_val is not None else ' (no solicitado)', (None if g2_val is not None else 'muted',))
        self.txt_resumen.insert('end', '  −  Salidas ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(s_val) if s_val is not None else ' (no solicitado)', (None if s_val is not None else 'muted',))
        self.txt_resumen.insert('end', '  =  ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(caja_calc) if caja_calc is not None else ' —', ('caja',))
        self.txt_resumen.insert('end', '\n\n')

        # 3) Crecimiento de la cartera = Cartera(desde) − Cartera(hasta)
        if 'Cartera en calle' in sel:
            car_desde = data.get('cartera_en_calle_desde')
            # Intentar obtener 'cartera_en_calle_hasta', si no existe (None), usar 'cartera_en_calle'
            car_hasta = data.get('cartera_en_calle_hasta')
            if car_hasta is None:
                car_hasta = data.get('cartera_en_calle')
        else:
            car_desde = None
            car_hasta = None
            
        crecimiento = (float(car_hasta) - float(car_desde)) if (car_desde is not None and car_hasta is not None) else None
        pct = None
        if crecimiento is not None and car_desde not in (None, 0):
            try:
                pct = (crecimiento / float(car_desde)) * 100
            except Exception:
                pct = None

        self.txt_resumen.insert('end', 'Crecimiento de la cartera: ', ('cartera',))
        self.txt_resumen.insert('end', '\n')
        self.txt_resumen.insert('end', '  Cartera (hasta) ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(car_hasta) if car_hasta is not None else ' (no solicitado)', (None if car_hasta is not None else 'muted',))
        self.txt_resumen.insert('end', '  −  Cartera (desde) ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(car_desde) if car_desde is not None else ' (no solicitado)', (None if car_desde is not None else 'muted',))
        self.txt_resumen.insert('end', '  =  ', ('mono',))
        self.txt_resumen.insert('end', fmt_money(crecimiento) if crecimiento is not None else ' —', ('cartera',))
        self.txt_resumen.insert('end', '\n')
        if pct is not None:
            signo = '+' if pct >= 0 else ''
            self.txt_resumen.insert('end', f"  Variación porcentual: {signo}{pct:.2f}%", ('cartera',))
        else:
            self.txt_resumen.insert('end', '  Variación porcentual: (no disponible)', ('muted',))

        # Actualizar labels nuevos y ganancia
        if hasattr(self, 'lbl_efectivo'):
            val = data.get('total_efectivo', 0)
            self.lbl_efectivo.config(text=f"$ {float(val):,.0f}" if val is not None else "$ 0")
        if hasattr(self, 'lbl_clavos'):
            val = data.get('total_clavos', 0)
            self.lbl_clavos.config(text=f"$ {float(val):,.0f}" if val is not None else "$ 0")
        if hasattr(self, 'lbl_ganancia'):
            val = data.get('ganancia', 0)
            self.lbl_ganancia.config(text=f"$ {float(val):,.0f}" if val is not None else "$ 0")

    def _compute_cartera_fallback_async(self, empleado_id: str, fecha_hasta: date) -> None:
        t0 = _pc()
        val = self._fallback_cartera_en_calle(empleado_id, fecha_hasta)
        elapsed = _pc() - t0
        def _apply():
            if val is None:
                self._cartera_fb_key = None
                return
            # Actualizar caché y UI con nuevo valor
            if isinstance(self._cached_metrics, dict):
                self._cached_metrics['cartera_en_calle'] = float(val)
                seleccion = self.metrics_combo.get_selected()
                if not seleccion or ('Todo' in seleccion):
                    seleccion = ['Valor cobrado','Préstamos','Intereses','Gastos','Entradas','Base','Ganancia','Salidas','Caja','Cartera en calle','Número de abonos']
                self._render_outputs(self._cached_metrics, seleccion)
                self._cartera_fb_key = None
        try:
            self.after(0, _apply)
        except Exception:
            # Si no hay loop UI activo
            _apply()


